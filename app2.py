import time

from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from googletrans import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

from env import URL, DriverLocation


def counter():
    result = driver.find_element(By.CLASS_NAME,'jANrlb').find_element(By.CLASS_NAME,'fontBodySmall').text
    result = result.replace(',', '')
    result = "".join(result.split())
    result = result.split(':')
    # result = result[0].split(':')
    return int(int(result[1])/10)+1


def scrolling(counter):
    print('scrolling...')
    scrollable_div = driver.find_element("xpath",
        '//div[@class="lXJj5c Hk4XGb "]')
    for _i in range(counter):
        scrolling = driver.execute_script(
            'document.getElementsByClassName("dS8AEf")[0].scrollTop = document.getElementsByClassName("dS8AEf")[0].scrollHeight',
            scrollable_div
        )
        time.sleep(3)


def get_data(driver):
    """
    this function get main text, score, name
    """
    print('get data...')
    
    # elements = driver.find_elements("xpath",
    #     '//div[@class="jftiEf fontBodyMedium "]')
    # driver.find_elements(By.XPATH, '//button[text()="Some text"]')
    # driver.find_elements(By.CLASS_NAME, "tomatoes")
    # elements = driver.find_elements(By.XPATH, '//div[@class="jftiEf fontBodyMedium "]')
    # elements = driver.find_elements(By.CLASS_NAME, "jftiEf fontBodyMedium")
    driver.execute_script("document.querySelector('input').blur();")
    button_element = driver.find_element(
        By.XPATH,
        "//button[@role='tab' and @aria-label='Reviews for AT Locksmiths London']")
    print("button", button_element)
    button_element.click()
    driver.implicitly_wait(10)
    
    elements = driver.find_elements(By.XPATH, '//div[@class="jftiEf fontBodyMedium "]')
    # print("elements = ", elements)
    lst_data = []
    translator = Translator()
    print("elements - jftiEf fontBodyMedium", len(elements))
    for data in elements:
        try:
            name_element = data.find_element(By.CLASS_NAME,
                'd4r55')
            name = name_element.text
        except NoSuchElementException:
            name = ""
        print("name = ", name)
        try:
            text_element = data.find_element(By.CLASS_NAME,
             'RfnDt')
            text = text_element.text
            reviews_number = ''.join(filter(str.isdigit, text))
        except NoSuchElementException:
            reviews_number = 0
        print("reviews_number = ", reviews_number)
        try:
            comment_element = data.find_element(By.CLASS_NAME,
                'wiI7pd')
            comment = comment_element.text
        except NoSuchElementException:
            comment = ""
        print("comment = ", comment)
        # comment = comment_element.text if comment_element else ""
        # translated_comment = translator.translate(" Название: Разочарован отсутствием улучшений в Google", dest='en')
        try:
            score_element = data.find_element(By.CLASS_NAME,
            'kvMYJc')
            score = score_element.get_attribute("aria-label")
            stars_number = ''.join(filter(str.isdigit, score))
        except NoSuchElementException:
            stars_number = 0
        print("stars_number = ", stars_number)

        lst_data.append([name + " from GoogleMaps", str(reviews_number) + " reviews", comment, str(stars_number) + " stars"])

    return lst_data


def write_to_xlsx(data):
    print('write to excel...')
    # cols = ["name", "comment", 'rating']
    cols = ["name", "reviews", "comments", 'rating']
    df = pd.DataFrame(data, columns=cols)
    df.to_excel('out2.xlsx')
    time.sleep(5)
    workbook = load_workbook('out2.xlsx')
    sheet = workbook.active
    new_data = ['Value1', 'Value2', 'Value3']
    max_row = sheet.max_row
    next_row = max_row + 1
    for i, value in enumerate(new_data):
        sheet.cell(row=next_row, column=i+1).value = value
    workbook.save('out2.xlsx')


if __name__ == "__main__":

    print('starting...')
    # options = webdriver.ChromeOptions()
    options = Options()    
    options.add_argument("--headless")  # show browser or not
    options.add_argument("--lang=en-US")
    options.add_experimental_option('prefs', {'intl.accept_languages': 'en,en_US'})
    options.add_argument('--ignore-certificate-errors')
    DriverPath = DriverLocation
    driver = webdriver.Chrome(options)
    # driver.execute_script("document.documentElement.lang = 'en';")

    driver.get(URL)
    driver.implicitly_wait(5)

    # counter = counter()
    # print(counter)
    # scrolling(counter)

    data = get_data(driver)
    driver.close()

    write_to_xlsx(data)
    print('Done!')
